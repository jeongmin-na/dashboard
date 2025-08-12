#!/usr/bin/env python3
"""
Cursor Admin API 프록시 서버
CORS 문제를 해결하고 실제 API 호출을 중계합니다.
"""

import http.server
import socketserver
import urllib.request
import urllib.parse
import json
import base64
from urllib.error import HTTPError, URLError
import ssl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

class CursorAPIProxy(http.server.SimpleHTTPRequestHandler):
    """Cursor Admin API 프록시 핸들러"""
    
    def __init__(self, *args, **kwargs):
        self.api_key = "key_e46368ce482125bbd568b7d55090c657e30e4b73c824f522cbc9ef9b1bf3f0d3"
        self.base_url = "https://api.cursor.com"
        super().__init__(*args, **kwargs)
    
    def do_OPTIONS(self):
        """CORS preflight 요청 처리"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        self.send_header('Access-Control-Max-Age', '86400')
        self.end_headers()
    
    def do_GET(self):
        """GET 요청 처리"""
        if self.path.startswith('/teams/'):
            self.handle_api_request('GET')
        else:
            # 정적 파일 서빙
            super().do_GET()
    
    def do_POST(self):
        """POST 요청 처리"""
        if self.path.startswith('/teams/'):
            self.handle_api_request('POST')
        elif self.path == '/send-email':
            self.handle_email_send()
        elif self.path == '/generate-xlsx':
            self.handle_generate_xlsx()
        else:
            self.send_error(404, "Not Found")
    
    def handle_api_request(self, method):
        """API 요청 처리"""
        try:
            # 요청 본문 읽기
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = None
            if content_length > 0:
                post_data = self.rfile.read(content_length)
            
            # API 요청 생성
            url = f"{self.base_url}{self.path}"
            
            # Basic Auth 설정
            credentials = f"{self.api_key}:"
            encoded_credentials = base64.b64encode(credentials.encode()).decode()
            
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Basic {encoded_credentials}'
            }
            
            # 요청 생성
            if method == 'GET':
                req = urllib.request.Request(url, headers=headers)
            else:  # POST
                req = urllib.request.Request(url, data=post_data, headers=headers, method='POST')
            
            print(f"프록시 요청: {method} {url}")
            if post_data:
                print(f"요청 데이터: {post_data.decode()}")
            
            # API 호출
            with urllib.request.urlopen(req) as response:
                response_data = response.read()
                response_headers = response.headers
                
                print(f"API 응답: {response.status}")
                print(f"응답 데이터: {response_data.decode()[:200]}...")
                
                # CORS 헤더 추가
                self.send_response(response.status)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
                
                # 원본 응답 헤더 복사
                for header, value in response_headers.items():
                    if header.lower() not in ['transfer-encoding', 'connection']:
                        self.send_header(header, value)
                
                self.end_headers()
                self.wfile.write(response_data)
                
        except HTTPError as e:
            print(f"HTTP 에러: {e.code} - {e.reason}")
            error_response = {
                'error': 'HTTP Error',
                'code': e.code,
                'message': e.reason
            }
            self.send_error_response(500, json.dumps(error_response))
            
        except URLError as e:
            print(f"URL 에러: {e.reason}")
            error_response = {
                'error': 'Connection Error',
                'message': str(e.reason)
            }
            self.send_error_response(500, json.dumps(error_response))
            
        except Exception as e:
            print(f"예상치 못한 에러: {e}")
            error_response = {
                'error': 'Internal Server Error',
                'message': str(e)
            }
            self.send_error_response(500, json.dumps(error_response))
    
    def send_error_response(self, status_code, error_data):
        """에러 응답 전송"""
        self.send_response(status_code)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(error_data.encode())
    
    def handle_email_send(self):
        """이메일 발송 처리"""
        try:
            # 요청 본문 읽기
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length > 0:
                post_data = self.rfile.read(content_length)
                email_data = json.loads(post_data.decode('utf-8'))
            else:
                raise ValueError("요청 데이터가 없습니다.")
            
            print(f"📧 이메일 발송 요청: {email_data}")
            
            # 이메일 발송
            result = self.send_email_via_smtp(email_data)
            
            # 성공 응답
            response_data = {
                'success': True,
                'message': '이메일이 성공적으로 발송되었습니다.',
                'sent_to': email_data.get('to_emails', []),
                'timestamp': email_data.get('timestamp', '')
            }
            
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(response_data, ensure_ascii=False).encode())
            
        except Exception as e:
            print(f"❌ 이메일 발송 실패: {e}")
            error_response = {
                'success': False,
                'error': '이메일 발송에 실패했습니다.',
                'message': str(e)
            }
            self.send_error_response(500, json.dumps(error_response, ensure_ascii=False))

    def handle_generate_xlsx(self):
        """프런트에서 보낸 시트 정의로 XLSX를 생성하여 바이너리로 응답"""
        try:
            if Workbook is None:
                raise RuntimeError('openpyxl 미설치')
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length <= 0:
                raise ValueError('요청 데이터가 없습니다.')
            payload = json.loads(self.rfile.read(content_length).decode('utf-8'))
            sheets = payload.get('sheets', [])
            filename = payload.get('filename', 'attachments.xlsx')

            wb = Workbook()
            # 기본 시트 제거
            default_ws = wb.active
            wb.remove(default_ws)
            for sheet in sheets:
                name = sheet.get('name', 'Sheet')[:31] or 'Sheet'
                ws = wb.create_sheet(title=name)
                headers = sheet.get('headers', [])
                rows = sheet.get('rows', [])
                if headers:
                    ws.append(headers)
                for row in rows:
                    ws.append(row)

            bio = BytesIO()
            wb.save(bio)
            data = bio.getvalue()

            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Content-Length', str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            print(f'❌ XLSX 생성 실패: {e}')
            self.send_error_response(500, json.dumps({'error':'XLSX 생성 실패','message':str(e)}, ensure_ascii=False))
    
    def send_email_via_smtp(self, email_data):
        """SMTP를 통한 이메일 발송"""
        # Gmail SMTP 설정
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        
        # 발신자 정보 (실제 Gmail 계정 정보로 변경 필요)
        sender_email = "jeongmin.na7@gmail.com"  # 실제 Gmail 주소로 변경 (기본 계정)
        sender_password = "rfwz pyja jtft igvh"   # Gmail 앱 비밀번호로 변경 (앱 비밀번호 생성 후 변경)

        # 프론트에서 전달된 발신자 표시 이름/이메일(옵션)
        requested_from_email = (email_data.get('from_email') or '').strip()
        requested_from_name = (email_data.get('from_name') or '').strip()
        
        print(f"📧 Gmail SMTP 설정 - 서버: {smtp_server}:{smtp_port}")
        print(f"📧 발신자: {sender_email}")
        
        # 수신자 이메일 목록
        to_emails = email_data.get('to_emails', [])
        if not to_emails:
            raise ValueError("수신자 이메일이 없습니다.")
        
        print(f"📬 수신자 목록: {to_emails}")
        
        # 이메일 메시지 생성
        msg = MIMEMultipart('mixed')
        display_name = requested_from_name if requested_from_name else 'Samsung AI Dashboard'
        from_header_email = requested_from_email if requested_from_email else sender_email
        msg['From'] = f"{display_name} <{from_header_email}>"
        # 수신자 회신 시 표시를 보장하기 위해 Reply-To를 명시적으로 설정
        msg['Reply-To'] = f"{display_name} <{from_header_email}>"
        msg['Subject'] = email_data.get('subject', '[Samsung AI Dashboard] 리포트')
        
        # HTML 본문 생성
        html_content = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .header {{ background-color: #4a9eff; color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 20px; }}
                .footer {{ background-color: #f5f5f5; padding: 15px; text-align: center; font-size: 12px; color: #666; }}
                .highlight {{ background-color: #fff3cd; padding: 10px; border-left: 4px solid #ffc107; margin: 10px 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 Samsung AI Dashboard 리포트</h1>
            </div>
            
            <div class="content">
                <h2>📝 메시지</h2>
                <p>{email_data.get('message', '메시지가 없습니다.')}</p>
            </div>
            
            <div class="footer">
                <p>이 이메일은 Samsung AI Dashboard에서 자동으로 발송되었습니다.</p>
                <p>© 2025 Samsung AI Experience Group</p>
            </div>
        </body>
        </html>
        """
        
        # 텍스트 본문 (HTML을 지원하지 않는 클라이언트용)
        text_content = f"""
        Samsung AI Dashboard 리포트

        메시지:
        {email_data.get('message', '메시지가 없습니다.')}
        ---
        이 이메일은 Samsung AI Dashboard에서 자동으로 발송되었습니다.
        © 2025 Samsung AI Experience Group
        """
        
        # 본문 첨부 (alternative 파트 생성)
        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText(text_content, 'plain', 'utf-8'))
        alt.attach(MIMEText(html_content, 'html', 'utf-8'))
        msg.attach(alt)

        # 첨부 파일: XLSX 시트 데이터가 포함된 경우 생성하여 첨부
        try:
            sheets_payload = email_data.get('attachmentsSheets')
            if sheets_payload and Workbook is not None:
                wb = Workbook()
                default_ws = wb.active
                wb.remove(default_ws)
                for sheet in sheets_payload.get('sheets', []):
                    name = sheet.get('name', 'Sheet')[:31] or 'Sheet'
                    ws = wb.create_sheet(title=name)
                    headers = sheet.get('headers', [])
                    rows = sheet.get('rows', [])
                    if headers:
                        ws.append(headers)
                    for row in rows:
                        ws.append(row)
                bio = BytesIO()
                wb.save(bio)
                xlsx_bytes = bio.getvalue()
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(xlsx_bytes)
                encoders.encode_base64(part)
                filename = sheets_payload.get('filename', 'attachments.xlsx')
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
        except Exception as e:
            print(f'⚠️ XLSX 첨부 생성 건너뜀: {e}')
        
        # SMTP 서버 연결 및 발송
        try:
            print(f"🔄 SMTP 서버 연결 시도: {smtp_server}:{smtp_port}")
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                print("✅ SMTP 서버 연결 성공")
                
                print("🔐 TLS 암호화 시작")
                server.starttls()
                
                print(f"🔑 Gmail 로그인 시도: {sender_email}")
                server.login(sender_email, sender_password)
                print("✅ Gmail 로그인 성공")
                
                # 각 수신자에게 개별 발송 (BCC 효과)
                for i, recipient in enumerate(to_emails):
                    # 개별 메시지 생성 (헤더 중복 방지)
                    individual_msg = MIMEMultipart('mixed')
                    individual_msg['From'] = f"{display_name} <{from_header_email}>"
                    individual_msg['To'] = recipient
                    individual_msg['Subject'] = email_data.get('subject', '[Samsung AI Dashboard] 리포트')
                    individual_msg['Reply-To'] = f"{display_name} <{from_header_email}>"
                    # 본문 첨부
                    alt2 = MIMEMultipart('alternative')
                    alt2.attach(MIMEText(text_content, 'plain', 'utf-8'))
                    alt2.attach(MIMEText(html_content, 'html', 'utf-8'))
                    individual_msg.attach(alt2)
                    # 동일 첨부 추가
                    for att in msg.get_payload()[1:]:  # alt 이후 첨부들
                        individual_msg.attach(att)
                    
                    print(f"📤 이메일 발송 중 ({i+1}/{len(to_emails)}): {recipient}")
                    server.send_message(individual_msg)
                    print(f"✅ 이메일 발송 완료: {recipient}")
                
                print(f"🎉 전체 이메일 발송 완료: {len(to_emails)}명")
                return True
                
        except smtplib.SMTPAuthenticationError as e:
            print(f"❌ Gmail 인증 실패: {e}")
            raise Exception("Gmail 인증 실패. 앱 비밀번호를 확인해주세요. Gmail 계정에서 2단계 인증을 활성화하고 앱 비밀번호를 생성해야 합니다.")
        except smtplib.SMTPRecipientsRefused as e:
            print(f"❌ 수신자 이메일 오류: {e}")
            raise Exception(f"수신자 이메일 오류: {e}")
        except smtplib.SMTPServerDisconnected as e:
            print(f"❌ SMTP 서버 연결 끊김: {e}")
            raise Exception("SMTP 서버 연결이 끊어졌습니다.")
        except smtplib.SMTPConnectError as e:
            print(f"❌ SMTP 서버 연결 실패: {e}")
            raise Exception(f"SMTP 서버 연결 실패: {e}")
        except smtplib.SMTPException as e:
            print(f"❌ SMTP 일반 오류: {e}")
            raise Exception(f"SMTP 오류: {e}")
        except Exception as e:
            print(f"❌ 예상치 못한 오류: {e}")
            raise Exception(f"이메일 발송 중 오류 발생: {str(e)}")

def run_proxy_server(port=8001):
    """프록시 서버 실행"""
    with socketserver.TCPServer(("", port), CursorAPIProxy) as httpd:
        print(f"🚀 Cursor API 프록시 서버가 포트 {port}에서 실행 중입니다...")
        print(f"📊 대시보드 접속: http://localhost:{port}/dash.html")
        print("🛑 서버를 중지하려면 Ctrl+C를 누르세요.")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n🛑 서버를 중지합니다...")
            httpd.shutdown()

if __name__ == "__main__":
    run_proxy_server() 
